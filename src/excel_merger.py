from pathlib import Path
from openpyxl import Workbook, load_workbook
import typer


class ExcelMerger:
    def __init__(self, folder_path: str, output_path: str = "./merged.pdf", overwrite: bool = False):
        self.folder_path = Path(folder_path)
        self.output_path = Path(output_path)
        self.overwrite = overwrite
        self._merged_workbook = None


    def merge_files(self):
        """Merge multiple Excel files into a single Workbook.

        - folder_path: path to folder containing .xlsx files
        - Returns: openpyxl.Workbook containing merged sheets

        Schema validation: all workbooks must have the same number of worksheets,
        the same sheet titles in the same order, and identical header (first row)
        values for each corresponding sheet. Otherwise raises ValueError.
        """
        if not self.folder_path.is_dir():
            raise ValueError(f"Provided folder path '{self.folder_path}' is not a valid directory")

        file_paths = self._list_excel_files(self.folder_path)
        if not file_paths:
            raise ValueError("No .xlsx files found in the specified folder")
        
        if(self.output_path.exists() and not self.overwrite):
            typer.confirm(f"File {self.output_path} already exists. Do you want to overwrite it?", abort=True)

        workbooks = [load_workbook(path) for path in file_paths]

        base_titles, base_headers, n_sheets = self._validate_workbooks(workbooks)

        merged_wb = Workbook()
        try:
            merged_wb.remove(merged_wb.active) # type: ignore
        except Exception:
            pass

        for title in base_titles:
            merged_wb.create_sheet(title=title)

        # Append rows: keep header from first workbook, skip headers for subsequent
        for idx, title in enumerate(base_titles):
            target_ws = merged_wb[title]
            first_wb = True
            for wb in workbooks:
                src_ws = wb.worksheets[idx]
                for jdx, row in enumerate(src_ws.iter_rows(values_only=True)):
                    if jdx == 0:
                        if first_wb:
                            target_ws.append(list(row))
                        # else: skip header row
                    else:
                        target_ws.append(list(row))
                first_wb = False

        self._merged_workbook = merged_wb

        return self

    def save(self):

        if self._merged_workbook is None:
            raise ValueError("No merged workbook to save. Please run merge_files() first.")

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.output_path.unlink(missing_ok=True)

        self._merged_workbook.save(self.output_path)

        print(f"Merged file saved as {self.output_path}")

    def format(self, formatter):
        """Apply formatting to the merged workbook using a Formatter instance."""
        if self._merged_workbook is None:
            raise ValueError("No merged workbook to format. Please run merge_files() first.")
        
        for ws in self._merged_workbook.worksheets:
            formatter.apply(ws)

        return self
    
    def _validate_workbooks(self, workbooks):

        # Validate same number of sheets
        sheet_counts = [len(wb.worksheets) for wb in workbooks]
        if len(set(sheet_counts)) != 1:
            raise ValueError(f"Workbooks have different number of worksheets: {sheet_counts}")
        n_sheets = sheet_counts[0]

        # Validate sheet titles (order matters)
        base_titles = [ws.title for ws in workbooks[0].worksheets]
        for idx, wb in enumerate(workbooks[1:], start=1):
            titles = [ws.title for ws in wb.worksheets]
            if titles != base_titles:
                raise ValueError( f"Workbook at index {idx} has different sheet titles: {titles} != {base_titles}" )

        # Validate headers for each sheet
        base_headers = []
        for s_idx in range(n_sheets):
            ws0 = workbooks[0].worksheets[s_idx]
            base_headers.append(self._get_header(ws0))

        for idx, wb in enumerate(workbooks[1:], start=1):
            for s_idx in range(n_sheets):
                ws = wb.worksheets[s_idx]
                headers = self._get_header(ws)
                if headers != base_headers[s_idx]:
                    raise ValueError( f"Workbook at index {idx}, sheet '{ws.title}' header mismatch: {headers} != {base_headers[s_idx]}" )
                
        return [base_titles, base_headers, n_sheets]
    
    def _get_header(self, ws):
        try:
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            return list(row)
        except StopIteration:
            return []

    def _list_excel_files(self, folder_path: Path):
        """Helper method to list .xlsx files in a folder."""
        return [f for f in folder_path.glob("*.xlsx")]