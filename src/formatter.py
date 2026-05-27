
from enum import Enum

from openpyxl.styles import Border, Font, PatternFill, Side
import yaml, re


class Formatter:
    class BorderStyle(Enum):
        MEDIUM_DASHED = 'mediumDashed'
        MEDIUM_DASH_DOT_DOT = 'mediumDashDotDot'
        DASH_DOT = 'dashDot'
        DASHED = 'dashed'
        SLANT_DASH_DOT = 'slantDashDot'
        DASH_DOT_DOT = 'dashDotDot'
        THICK = 'thick'
        THIN = 'thin'
        DOTTED = 'dotted'
        DOUBLE = 'double'
        MEDIUM = 'medium'
        HAIR = 'hair'
        MEDIUM_DASH_DOT = 'mediumDashDot'

    def __init__(self, config_path):
        with open(config_path, 'r') as f:
            self.config = yaml.safe_load(f)

    def apply(self, worksheet):
        for idx, row in enumerate(worksheet.iter_rows()):
            if idx == 0:
                self._apply_style(self.config.get('header', {}), row)
            else:
                self._apply_style(self.config.get('cell', {}), row)

    def _apply_style(self, style, row):
        for cell in row:
            if 'bg-color' in style:
                cell.fill = PatternFill(fill_type='solid', start_color=self._parse_color(style['bg-color']))
            if 'text-color' in style:
                cell.font = cell.font.copy(color=self._parse_color(style['text-color']))
            if 'font-size' in style:
                cell.font = cell.font.copy(size=style['font-size'])
            if 'text-bold' in style:
                cell.font = cell.font.copy(bold=style['text-bold'])
            if 'text-italic' in style:
                cell.font = cell.font.copy(italic=style['text-italic'])
            if 'text-strikethrough' in style:
                cell.font = cell.font.copy(strike=style['text-strikethrough'])
            if 'font-name' in style:
                cell.font = Font(name=style['font-name'], size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, strike=cell.font.strike, color=cell.font.color)
            if 'border' in style:
                cell.border = self._create_border(style['border'])

    def _create_border(self, border_str):
        border_styles = border_str.split()

        if len(border_styles) != 2:
            raise ValueError("Invalid border format. Please use 'border_type color' format, e.g. 'thick #dddddd'.")
        
        border_type, color = border_styles

        if border_type not in [style.value for style in self.BorderStyle]:
            raise ValueError("Invalid border type. Supported types are: " + ", ".join([style.value for style in self.BorderStyle]))

        return Border(
            left=Side(style=border_type, color=self._parse_color(color)),
            right=Side(style=border_type, color=self._parse_color(color)),
            top=Side(style=border_type, color=self._parse_color(color)),
            bottom=Side(style=border_type, color=self._parse_color(color))
        )

    def _parse_color(self, color_str):
        if re.match(r'^#[a-fA-F0-9]{8}$', color_str):
            return color_str.strip('#')
        else:
            raise ValueError(f"Invalid color format: {color_str}. Please use hex format like '#FFRRGGBB'.")
        